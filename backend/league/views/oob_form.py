"""
GET /api/forms/oob/

Returns a filled Out-of-Boundary Waiver Request Form (.xlsx) with WTLL's
league name, league ID, and League President name/email pre-populated from
the database.  All other fields (player info, home league, decisions) are
left blank for the submitter to fill in.
"""
import io
import os

from django.http import HttpResponse
from rest_framework.views import APIView

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "form_templates",
    "OOB_template.xlsx",
)


class OOBFormView(APIView):
    """GET /api/forms/oob/ — download a pre-filled OOB waiver form."""

    authentication_classes = []  # accessible to any logged-in user
    permission_classes = []

    def get(self, request):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return HttpResponse("openpyxl not installed", status=500)

        # ── Pull WTLL identity ───────────────────────────────────────────────
        league_name  = "Washington Township Little League"
        league_id    = "1140814"
        try:
            from league.models.site_settings import LeagueIdentity
            li = LeagueIdentity.get()
            league_name = li.league_name or league_name
            league_id   = getattr(li, "little_league_id", None) or league_id
        except Exception:
            pass

        # Format the ID with leading # if not already present
        if league_id and not league_id.startswith("#"):
            league_id = f"#{league_id}"

        # ── Pull League President from board ─────────────────────────────────
        president_name  = ""
        president_email = ""
        try:
            from league.models.board_member import BoardMember
            president = BoardMember.objects.filter(
                role__iexact="president", is_active=True
            ).first()
            if president:
                president_name  = president.full_name
                president_email = president.email or ""
        except Exception:
            pass

        # ── Load template and fill placeholders ──────────────────────────────
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        substitutions = {
            "{{LEAGUE_NAME}}":     league_name,
            "{{LEAGUE_ID}}":       league_id,
            "{{PRESIDENT_NAME}}":  president_name,
            "{{PRESIDENT_EMAIL}}": president_email,
        }

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in substitutions:
                    cell.value = substitutions[cell.value]

        # ── Stream to response ───────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="OOB_Waiver_Request_Form.xlsx"'
        return response
