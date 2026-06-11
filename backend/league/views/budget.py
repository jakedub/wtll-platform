"""
Budget views: line-item CRUD, summary/gap analysis, approval, XLSX export.
"""
import io
from decimal import Decimal

from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from league.models import BudgetLine, BudgetApproval
from league.models.budget import BUDGET_CATEGORIES
from league.serializers.budget_serializer import BudgetLineSerializer, BudgetApprovalSerializer


# ── Line items ────────────────────────────────────────────────────────────────


class BudgetLineListCreateView(APIView):
    """
    GET  /api/budget/lines/?year=<year>
    POST /api/budget/lines/
    """

    def get(self, request):
        qs = BudgetLine.objects.all()
        year = request.query_params.get("year")
        if year:
            qs = qs.filter(year=year)
        return Response(BudgetLineSerializer(qs, many=True).data)

    def post(self, request):
        serializer = BudgetLineSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=400)


class BudgetLineDetailView(APIView):
    """PATCH / DELETE /api/budget/lines/<pk>/"""

    def _get(self, pk):
        try:
            return BudgetLine.objects.get(pk=pk)
        except BudgetLine.DoesNotExist:
            return None

    def patch(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({"error": "Not found."}, status=404)
        # If estimate is being set explicitly, mark as override
        if "estimate" in request.data and request.data.get("estimate") is not None:
            request.data["estimate_override"] = True
        serializer = BudgetLineSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def delete(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({"error": "Not found."}, status=404)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Bulk reorder ──────────────────────────────────────────────────────────────

class BudgetLineReorderView(APIView):
    """
    POST /api/budget/lines/reorder/
    Body: { items: [ {id, sort_order}, ... ] }
    """

    def post(self, request):
        items = request.data.get("items", [])
        for item in items:
            BudgetLine.objects.filter(pk=item["id"]).update(sort_order=item["sort_order"])
        return Response({"ok": True})


# ── Summary / gap analysis ────────────────────────────────────────────────────

class BudgetSummaryView(APIView):
    """
    GET /api/budget/summary/?year=<year>
    Returns totals by category plus overall gap analysis.
    """

    def get(self, request):
        year = request.query_params.get("year")
        qs = BudgetLine.objects.all()
        if year:
            qs = qs.filter(year=year)

        expense_lines = qs.filter(is_revenue=False)
        revenue_lines = qs.filter(is_revenue=True)

        def total(lines, field="actual"):
            return float(sum(
                (getattr(ln, field) or 0) for ln in lines
            ))

        def eff_total(lines):
            return float(sum(
                (ln.effective_estimate or 0) for ln in lines
            ))

        # Per-category expense totals
        by_cat = {}
        for code, label in BUDGET_CATEGORIES:
            if code == "REVENUE":
                continue
            cat_lines = [ln for ln in expense_lines if ln.category == code]
            by_cat[code] = {
                "label": label,
                "actual": total(cat_lines, "actual"),
                "estimate": eff_total(cat_lines),
            }

        total_expense_actual = total(expense_lines, "actual")
        total_expense_est = eff_total(expense_lines)
        total_revenue_actual = total(revenue_lines, "actual")
        total_revenue_est = eff_total(revenue_lines)

        # Approval
        approval = None
        if year:
            appr = BudgetApproval.objects.filter(year=year).first()
            if appr:
                approval = BudgetApprovalSerializer(appr).data

        return Response({
            "year": year,
            "by_category": by_cat,
            "total_expenses": {
                "actual": total_expense_actual,
                "estimate": total_expense_est,
            },
            "total_revenue": {
                "actual": total_revenue_actual,
                "estimate": total_revenue_est,
            },
            "gap": {
                "actual": total_revenue_actual - total_expense_actual,
                "estimate": total_revenue_est - total_expense_est,
            },
            "approval": approval,
        })


# ── Approval ──────────────────────────────────────────────────────────────────

class BudgetApprovalView(APIView):
    """
    POST /api/budget/approve/    { year, approved_by, notes }
    GET  /api/budget/approve/?year=<year>
    DELETE /api/budget/approve/<year>/   — revoke approval
    """

    def get(self, request):
        year = request.query_params.get("year")
        qs = BudgetApproval.objects.all()
        if year:
            qs = qs.filter(year=year)
        return Response(BudgetApprovalSerializer(qs, many=True).data)

    def post(self, request):
        year = request.data.get("year")
        if not year:
            return Response({"error": "year required."}, status=400)
        # Upsert: delete existing approval for year then create new
        BudgetApproval.objects.filter(year=year).delete()
        serializer = BudgetApprovalSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=400)

    def delete(self, request, year=None):
        deleted, _ = BudgetApproval.objects.filter(year=year).delete()
        if not deleted:
            return Response({"error": "No approval found."}, status=404)
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── XLSX Export ───────────────────────────────────────────────────────────────

class BudgetExportView(APIView):
    """GET /api/budget/export/?year=<year>"""

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response({"error": "openpyxl not installed."}, status=500)

        year = request.query_params.get("year", "")
        qs = BudgetLine.objects.all()
        if year:
            qs = qs.filter(year=year)

        RED_FILL   = PatternFill("solid", start_color="C41230")
        CAT_FILL   = PatternFill("solid", start_color="F0F0F0")
        HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        CAT_FONT    = Font(name="Arial", bold=True, size=10)
        BODY_FONT   = Font(name="Arial", size=10)
        CENTER = Alignment(horizontal="center", vertical="center")
        LEFT   = Alignment(horizontal="left",   vertical="center")
        RIGHT  = Alignment(horizontal="right",  vertical="center")
        thin   = Side(style="thin", color="D0D0D0")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Budget {year}" if year else "Budget"

        # Title
        ws.merge_cells("A1:F1")
        tc = ws["A1"]
        tc.value = f"WTLL Budget — {year}" if year else "WTLL Budget"
        tc.font = Font(name="Arial", bold=True, size=14)
        tc.alignment = CENTER
        ws.row_dimensions[1].height = 28

        # Column headers
        headers = ["Category", "Item", "Owner", "Prior Year Actual", "Est. (Actual+5%)", "Override Estimate"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = RED_FILL
            cell.alignment = CENTER
            cell.border = BORDER
        ws.row_dimensions[2].height = 16

        row = 3
        current_cat = None
        expense_actual = Decimal("0")
        expense_est = Decimal("0")
        revenue_actual = Decimal("0")
        revenue_est = Decimal("0")

        for ln in qs:
            # Category separator row
            if ln.category != current_cat:
                current_cat = ln.category
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                cell = ws.cell(row=row, column=1, value=ln.get_category_display())
                cell.font = CAT_FONT
                cell.fill = CAT_FILL
                cell.border = BORDER
                ws.row_dimensions[row].height = 15
                row += 1

            eff = ln.effective_estimate or Decimal("0")
            act = ln.actual or Decimal("0")
            if ln.is_revenue:
                revenue_actual += act
                revenue_est += eff
            else:
                expense_actual += act
                expense_est += eff

            values = [
                ln.get_category_display(),
                ln.item,
                ln.owner_role,
                float(act),
                float(ln.actual * Decimal("1.05")) if ln.actual and not ln.estimate_override else "",
                float(ln.estimate) if ln.estimate_override and ln.estimate else "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = RIGHT if col >= 4 else LEFT
                if col >= 4 and isinstance(val, float):
                    cell.number_format = '"$"#,##0.00'
            row += 1

        # Summary
        row += 1
        for label, act, est in [
            ("Total Expenses", float(expense_actual), float(expense_est)),
            ("Total Revenue",  float(revenue_actual), float(revenue_est)),
            ("Net (Revenue – Expenses)", float(revenue_actual - expense_actual), float(revenue_est - expense_est)),
        ]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = Font(name="Arial", bold=True, size=10)
            lc.border = BORDER
            for col, val in [(4, act), (5, est)]:
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Arial", bold=True, size=10, color="C41230" if label.startswith("Net") else "000000")
                c.border = BORDER
                c.alignment = RIGHT
                c.number_format = '"$"#,##0.00'
            row += 1

        col_widths = [18, 36, 28, 16, 16, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"wtll_budget_{year}.xlsx" if year else "wtll_budget.xlsx"
        resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


class BudgetYearsView(APIView):
    """GET /api/budget/years/ — returns all years that have budget lines."""

    def get(self, request):
        from league.models import BudgetLine
        years = list(
            BudgetLine.objects.values_list("year", flat=True)
            .distinct().order_by("-year")
        )
        return Response({"years": years})


class BudgetCopyYearView(APIView):
    """
    POST /api/budget/copy-year/
    Body: { from_year, to_year, multiplier }
    Copies all lines from from_year to to_year.
    New lines: actual = from_year effective_estimate, estimate = actual * multiplier.
    Skips if to_year already has lines.
    """

    def post(self, request):
        from league.models import BudgetLine
        from_year = request.data.get("from_year")
        to_year = request.data.get("to_year")
        multiplier = float(request.data.get("multiplier", 1.05))
        overwrite = bool(request.data.get("overwrite", False))

        if not from_year or not to_year:
            return Response({"error": "from_year and to_year required."}, status=400)
        if int(from_year) == int(to_year):
            return Response({"error": "from_year and to_year must be different."}, status=400)

        source = BudgetLine.objects.filter(year=from_year)
        if not source.exists():
            return Response({"error": f"No budget lines found for {from_year}."}, status=404)

        existing = BudgetLine.objects.filter(year=to_year)
        if existing.exists() and not overwrite:
            return Response(
                {"error": f"Year {to_year} already has {existing.count()} lines. Pass overwrite=true to replace."},
                status=409,
            )
        if overwrite:
            existing.delete()

        created = []
        for src in source.order_by("sort_order"):
            new_actual = src.effective_estimate or Decimal("0")
            if not isinstance(new_actual, Decimal):
                new_actual = Decimal(str(new_actual))
            new_estimate = (new_actual * Decimal(str(multiplier))).quantize(Decimal("0.01"))
            BudgetLine.objects.create(
                year=int(to_year),
                category=src.category,
                item=src.item,
                sub_group=src.sub_group,
                owner_role=src.owner_role,
                is_revenue=src.is_revenue,
                actual=new_actual,
                estimate=new_estimate,
                estimate_override=True,
                notes=src.notes,
                sort_order=src.sort_order,
            )
            created.append(src.item)

        return Response({"created": len(created), "to_year": to_year, "from_year": from_year}, status=201)
