"""
Draft views: CRUD, available players, draft/undraft players, team stats, state, export.
"""
import io
import traceback
from datetime import date

from django.db import transaction
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from league.models import Draft, DraftSelection, Player, Team, Division, Evaluation
from league.serializers.draft_serializer import DraftSerializer, DraftSelectionSerializer


# ── Draft CRUD ────────────────────────────────────────────────────────────────

class DraftListCreateView(APIView):
    """GET /api/drafts/  POST /api/drafts/"""

    def get(self, request):
        drafts = Draft.objects.select_related("division").order_by("-year", "name")
        return Response(DraftSerializer(drafts, many=True).data)

    def post(self, request):
        serializer = DraftSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DraftDetailView(APIView):
    """GET/PATCH/DELETE /api/drafts/<id>/"""

    def _get(self, pk):
        try:
            return Draft.objects.select_related("division").get(pk=pk)
        except Draft.DoesNotExist:
            return None

    def get(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({"error": "Not found."}, status=404)
        return Response(DraftSerializer(obj).data)

    def patch(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({"error": "Not found."}, status=404)
        serializer = DraftSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def delete(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({"error": "Not found."}, status=404)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Draft state (full snapshot) ───────────────────────────────────────────────

class DraftStateView(APIView):
    """
    GET /api/drafts/<id>/state/
    Returns the draft, all selected teams, and selections grouped by team.
    Includes evaluation data for each drafted player.
    """

    def get(self, request, pk):
        try:
            draft = Draft.objects.select_related("division").prefetch_related("selected_teams").get(pk=pk)
        except Draft.DoesNotExist:
            return Response({"error": "Not found."}, status=404)

        selections = DraftSelection.objects.filter(draft=draft).select_related("player", "team", "division")

        selections_by_team: dict = {}
        for sel in selections:
            tid = str(sel.team_id)
            if tid not in selections_by_team:
                selections_by_team[tid] = []
            player = sel.player
            evaluation = player.evaluations.filter(season_year=draft.year).first()
            selections_by_team[tid].append({
                "selection_id": sel.id,
                "id": player.id,
                "name": f"{player.first_name} {player.last_name}",
                "first_name": player.first_name,
                "last_name": player.last_name,
                "jersey_size": player.jersey_size or "",
                "batting_hand": player.batting_hand or "",
                "throwing_hand": player.throwing_hand or "",
                "sport": player.sport,
                "tier_spot": getattr(evaluation, "tier_spot", None),
                "overall_total": getattr(evaluation, "overall_total", None),
                "total_hitting": getattr(evaluation, "total_hitting", None),
                "total_fielding": getattr(evaluation, "total_fielding", None),
                "total_throwing": getattr(evaluation, "total_throwing", None),
                "total_pitching": getattr(evaluation, "total_pitching", None),
                "total_catcher": getattr(evaluation, "total_catcher", None),
                "is_pitcher": player.is_pitcher,
                "is_catcher": player.is_catcher,
            })

        selected_teams = [
            {"id": t.id, "name": t.name, "coach": t.coach or "", "assistant_coach": t.assistant_coach or "", "jersey_color": t.jersey_color or ""}
            for t in draft.selected_teams.all()
        ]

        return Response({
            "draft": DraftSerializer(draft).data,
            "selected_teams": selected_teams,
            "selections_by_team": selections_by_team,
        })


# ── Available players ─────────────────────────────────────────────────────────

class AvailablePlayersView(APIView):
    """
    GET /api/drafts/<id>/available-players/?division=<id>
    Returns eligible players not yet drafted in this draft, with eval data.
    """

    def get(self, request, pk):
        try:
            draft = Draft.objects.get(pk=pk)
        except Draft.DoesNotExist:
            return Response({"error": "Not found."}, status=404)

        division_id = request.query_params.get("division") or draft.division_id

        drafted_ids = DraftSelection.objects.filter(draft=draft).values_list("player_id", flat=True)

        # Derive sport from the division name so the filter is always correct
        # regardless of whether player.sport was set correctly on import.
        from league.models import Division as Div
        draft_sport = "baseball"
        try:
            div_obj = Div.objects.get(pk=division_id)
            if "softball" in (div_obj.name or "").lower():
                draft_sport = "softball"
        except Div.DoesNotExist:
            pass

        # Get all active players enrolled in this division matching the sport.
        # Eligibility is not required — ineligible players can still be drafted.
        # (Eligibility matters for All Stars paperwork, not regular season draft.)
        players = Player.objects.filter(
            is_active=True,
            is_archived=False,
            enrollments__division_id=division_id,
            sport__iexact=draft_sport,
        ).exclude(id__in=drafted_ids).distinct()

        data = []
        for p in players:
            evaluation = p.evaluations.filter(season_year=draft.year).first()
            data.append({
                "id": p.id,
                "name": f"{p.first_name} {p.last_name}",
                "first_name": p.first_name,
                "last_name": p.last_name,
                "batting_hand": p.batting_hand or "",
                "throwing_hand": p.throwing_hand or "",
                "jersey_size": p.jersey_size or "",
                "sport": p.sport,
                "tier_spot": getattr(evaluation, "tier_spot", None),
                "overall_total": getattr(evaluation, "overall_total", None),
                "total_hitting": getattr(evaluation, "total_hitting", None),
                "total_fielding": getattr(evaluation, "total_fielding", None),
                "total_throwing": getattr(evaluation, "total_throwing", None),
                "total_pitching": getattr(evaluation, "total_pitching", None),
                "total_catcher": getattr(evaluation, "total_catcher", None),
                "is_pitcher": p.is_pitcher,
                "is_catcher": p.is_catcher,
            })

        # Sort by tier_spot asc, then overall_total desc
        data.sort(key=lambda p: (p["tier_spot"] or 99, -(p["overall_total"] or 0)))
        return Response(data)


# ── Draft / undraft players ───────────────────────────────────────────────────

class DraftPlayerView(APIView):
    """
    POST   /api/drafts/<id>/pick/   { player_id, team_id }  — draft a player
    DELETE /api/drafts/<id>/pick/   { player_id, team_id }  — undo a pick
    """

    def post(self, request, pk):
        try:
            draft = Draft.objects.get(pk=pk)
        except Draft.DoesNotExist:
            return Response({"error": "Draft not found."}, status=404)

        player_id = request.data.get("player_id")
        team_id = request.data.get("team_id")

        if not player_id or not team_id:
            return Response({"error": "player_id and team_id required."}, status=400)

        try:
            player = Player.objects.get(id=player_id)
            team = Team.objects.get(id=team_id)
        except (Player.DoesNotExist, Team.DoesNotExist) as e:
            return Response({"error": str(e)}, status=404)

        if DraftSelection.objects.filter(draft=draft, player=player).exists():
            return Response({"error": f"{player.first_name} {player.last_name} is already drafted."}, status=400)

        division = team.division or draft.division
        selection = DraftSelection.objects.create(draft=draft, player=player, team=team, division=division)
        return Response(DraftSelectionSerializer(selection).data, status=status.HTTP_201_CREATED)

    def delete(self, request, pk):
        player_id = request.data.get("player_id")
        if not player_id:
            return Response({"error": "player_id required."}, status=400)

        deleted, _ = DraftSelection.objects.filter(draft_id=pk, player_id=player_id).delete()
        if deleted == 0:
            return Response({"error": "Selection not found."}, status=404)
        return Response({"deleted": True})


# ── Team management ───────────────────────────────────────────────────────────

class SaveDraftTeamsView(APIView):
    """POST /api/drafts/<id>/teams/  { team_ids: [int] }"""

    def post(self, request, pk):
        try:
            draft = Draft.objects.get(pk=pk)
        except Draft.DoesNotExist:
            return Response({"error": "Not found."}, status=404)

        team_ids = request.data.get("team_ids", [])
        draft.selected_teams.set(team_ids)
        return Response({"saved": team_ids})


class DraftTeamStatsView(APIView):
    """GET /api/drafts/<id>/team-stats/  — pitcher/catcher counts per team"""

    def get(self, request, pk):
        try:
            draft = Draft.objects.get(pk=pk)
        except Draft.DoesNotExist:
            return Response({"error": "Not found."}, status=404)

        team_ids = draft.selected_teams.values_list("id", flat=True)
        teams = Team.objects.filter(id__in=team_ids).order_by("name")
        stats = []

        for team in teams:
            selections = DraftSelection.objects.filter(draft=draft, team=team).select_related("player")
            count = selections.count()
            pitchers = catchers = 0
            overall_sum = 0

            for sel in selections:
                ev = sel.player.evaluations.filter(season_year=draft.year).first()
                if ev:
                    if (ev.total_pitching or 0) > 0:
                        pitchers += 1
                    if (ev.total_catcher or 0) > 0:
                        catchers += 1
                    overall_sum += ev.overall_total or 0

            stats.append({
                "team_id": team.id,
                "team_name": team.name,
                "coach": team.coach or "",
                "player_count": count,
                "pitchers": pitchers,
                "catchers": catchers,
                "avg_overall": round(overall_sum / count, 1) if count else 0,
            })

        return Response(stats)


class FallBallAutoAssignView(APIView):
    """
    POST /api/drafts/<id>/auto-assign/
    Fall Ball only. Assigns every unselected player to the single team
    in their enrolled division. Players already drafted are skipped.
    Returns a summary: { assigned, skipped_no_team, skipped_already_drafted }
    """

    def post(self, request, pk):
        try:
            draft = Draft.objects.select_related("division__program").get(pk=pk)
        except Draft.DoesNotExist:
            return Response({"error": "Draft not found."}, status=404)

        if not draft.division.program or draft.division.program.program_type != "FALL_BALL":
            return Response({"error": "Auto-assign is only available for Fall Ball drafts."}, status=400)

        from league.models.player_program_enrollment import PlayerProgramEnrollment

        # Players in this draft's division who haven't been drafted yet
        already_drafted_ids = set(
            DraftSelection.objects.filter(draft=draft).values_list("player_id", flat=True)
        )

        enrollments = PlayerProgramEnrollment.objects.select_related(
            "player", "division", "team"
        ).filter(
            division=draft.division,
            player__is_archived=False,
        ).exclude(player_id__in=already_drafted_ids)

        assigned = 0
        skipped_no_team = 0
        skipped_already_drafted = len(already_drafted_ids)

        with transaction.atomic():
            for enrollment in enrollments:
                if not enrollment.division:
                    skipped_no_team += 1
                    continue

                # Find the one team in this division for the draft year
                team = Team.objects.filter(
                    division=enrollment.division,
                    year=draft.year,
                    is_active=True,
                ).first()

                if not team:
                    skipped_no_team += 1
                    continue

                DraftSelection.objects.create(
                    draft=draft,
                    player=enrollment.player,
                    team=team,
                    division=enrollment.division,
                )
                # Also update the enrollment team
                enrollment.team = team
                enrollment.save(update_fields=["team"])
                assigned += 1

        return Response({
            "assigned": assigned,
            "skipped_no_team": skipped_no_team,
            "skipped_already_drafted": skipped_already_drafted,
            "message": (
                f"Auto-assigned {assigned} players. "
                f"{skipped_no_team} skipped (no team found). "
                f"{skipped_already_drafted} already drafted."
            ),
        })


class MarkDraftCompleteView(APIView):
    """POST /api/drafts/<id>/complete/"""

    def post(self, request, pk):
        try:
            draft = Draft.objects.get(pk=pk)
        except Draft.DoesNotExist:
            return Response({"error": "Not found."}, status=404)
        draft.is_complete = True
        draft.save()
        return Response(DraftSerializer(draft).data)


# ── XLSX Export ───────────────────────────────────────────────────────────────

class DraftExportView(APIView):
    """
    GET /api/drafts/<id>/export/
    Two-sheet XLSX: Draft Results + Jersey Roster Sheet.
    """

    def get(self, request, pk):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response({"error": "openpyxl not installed."}, status=500)

        try:
            draft = Draft.objects.select_related("division").get(pk=pk)
        except Draft.DoesNotExist:
            return Response({"error": "Not found."}, status=404)

        HEADER_FILL = PatternFill("solid", start_color="1F3864")
        HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        TEAM_FILL   = PatternFill("solid", start_color="D9E1F2")
        TEAM_FONT   = Font(name="Arial", bold=True, size=10)
        BODY_FONT   = Font(name="Arial", size=10)
        CENTER      = Alignment(horizontal="center", vertical="center")
        LEFT        = Alignment(horizontal="left", vertical="center")
        thin        = Side(style="thin", color="BFBFBF")
        BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
        JERSEY_SIZES = ["YXS", "YS", "YM", "YL", "YXL", "AS", "AM", "AL", "AXL", "AXXL"]

        def style_header(ws, row, cols):
            for c in range(1, cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = HEADER_FONT; cell.fill = HEADER_FILL
                cell.alignment = CENTER; cell.border = BORDER

        def style_team(ws, row, cols):
            for c in range(1, cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = TEAM_FONT; cell.fill = TEAM_FILL; cell.border = BORDER

        def style_body(ws, row, cols):
            for c in range(1, cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = BODY_FONT; cell.border = BORDER
                cell.alignment = LEFT if c == 1 else CENTER

        def set_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        teams = Team.objects.filter(
            id__in=DraftSelection.objects.filter(draft=draft).values_list("team_id", flat=True).distinct()
        ).order_by("name")

        wb = Workbook()

        # ── Sheet 1: Draft Results ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Draft Results"
        COLS1 = ["Player Name", "Batting Hand", "Throwing Hand", "Jersey Size",
                 "Tier", "Hitting", "Fielding", "Throwing", "Pitching", "Catcher", "Overall"]
        N1 = len(COLS1)

        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N1)
        tc = ws1.cell(row=1, column=1, value=f"{draft.name} — {draft.year} Draft Results")
        tc.font = Font(name="Arial", bold=True, size=13); tc.alignment = CENTER
        ws1.row_dimensions[1].height = 24

        r = 2
        for team in teams:
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N1)
            ws1.cell(row=r, column=1, value=f"{team.name}  |  Coach: {team.coach or '—'}  |  Asst: {team.assistant_coach or '—'}")
            style_team(ws1, r, N1); ws1.row_dimensions[r].height = 18; r += 1

            for c, label in enumerate(COLS1, 1):
                ws1.cell(row=r, column=c, value=label)
            style_header(ws1, r, N1); ws1.row_dimensions[r].height = 16; r += 1

            for sel in DraftSelection.objects.filter(draft=draft, team=team).select_related("player").order_by("selected_at"):
                p = sel.player
                ev = p.evaluations.filter(season_year=draft.year).first()
                row_vals = [
                    f"{p.first_name} {p.last_name}",
                    p.batting_hand or "", p.throwing_hand or "", p.jersey_size or "",
                    getattr(ev, "tier_spot", "") or "",
                    getattr(ev, "total_hitting", "") if ev else "",
                    getattr(ev, "total_fielding", "") if ev else "",
                    getattr(ev, "total_throwing", "") if ev else "",
                    getattr(ev, "total_pitching", "") if ev else "",
                    getattr(ev, "total_catcher", "") if ev else "",
                    getattr(ev, "overall_total", "") if ev else "",
                ]
                for c, val in enumerate(row_vals, 1):
                    ws1.cell(row=r, column=c, value=val)
                style_body(ws1, r, N1); r += 1
            r += 1

        set_widths(ws1, [26, 12, 13, 11, 6, 8, 8, 9, 9, 9, 9])

        # ── Sheet 2: Jersey Roster Sheet ───────────────────────────────────────
        ws2 = wb.create_sheet(title="Jersey Roster Sheet")
        N2 = 4 + len(JERSEY_SIZES)

        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N2)
        tc2 = ws2.cell(row=1, column=1, value=f"{draft.name} — {draft.year} Roster Sheet")
        tc2.font = Font(name="Arial", bold=True, size=13); tc2.alignment = CENTER
        ws2.row_dimensions[1].height = 24

        r = 2
        divisions = Division.objects.filter(
            id__in=teams.values_list("division_id", flat=True).distinct()
        ).order_by("name")

        for div in divisions:
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N2)
            dc = ws2.cell(row=r, column=1, value=f"Division: {div.name}")
            dc.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            dc.fill = PatternFill("solid", start_color="2E4057"); dc.alignment = LEFT
            ws2.row_dimensions[r].height = 18; r += 1

            # Header
            for c, h in enumerate(["Team Name", "Jersey Color", "Coach", "Asst Coach"] + JERSEY_SIZES, 1):
                ws2.cell(row=r, column=c, value=h)
            style_header(ws2, r, N2); ws2.row_dimensions[r].height = 16; r += 1

            for team in teams.filter(division_id=div.id):
                sizes = {s: 0 for s in JERSEY_SIZES}
                for sel in DraftSelection.objects.filter(draft=draft, team=team).select_related("player"):
                    sz = (sel.player.jersey_size or "").upper().strip()
                    if sz in sizes:
                        sizes[sz] += 1

                ws2.cell(row=r, column=1, value=team.name)
                ws2.cell(row=r, column=2, value=team.jersey_color or "—")
                ws2.cell(row=r, column=3, value=team.coach or "—")
                ws2.cell(row=r, column=4, value=team.assistant_coach or "—")
                for i, sz in enumerate(JERSEY_SIZES, 5):
                    cnt = sizes[sz]
                    ws2.cell(row=r, column=i, value=cnt if cnt > 0 else "")
                style_body(ws2, r, N2); r += 1
            r += 1

        set_widths(ws2, [22, 14, 20, 20] + [7] * len(JERSEY_SIZES))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"draft_{draft.name.replace(' ', '_')}_{draft.year}.xlsx"
        resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
