"""
Schedule Generator — round-robin schedule generation and SportsConnect xlsx export.

POST /api/schedules/generate/
  Accepts teams + date config, returns a list of game dicts.

POST /api/schedules/export/
  Accepts the finalized game list, returns a SportsConnect-compatible xlsx download.
"""
from datetime import date, time, datetime, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response


# ── Round-robin algorithm ─────────────────────────────────────────────────────

def _generate_base_rounds(teams: list) -> list:
    """
    Standard round-robin (circle method) with balanced home/away assignment.
    Returns a list of rounds; each round is a list of (away, home) tuples.
    Injects 'Bye' for odd team counts.

    The rotation keeps index 0 fixed and rotates indices 1..n-1 clockwise,
    which avoids the IndexError from the slice-based approach.
    Home/away is assigned greedily to the team with fewer home games so far.
    """
    t = list(teams)
    if len(t) % 2 == 1:
        t.append("Bye")

    n = len(t)
    if n < 2:
        return []

    # Track home counts for balanced assignment
    home_count = {team: 0 for team in t}

    # Slot indices: slots[i] is which team occupies position i this round
    slots = list(range(n))
    rounds = []

    for _ in range(n - 1):
        pairings = []
        for i in range(n // 2):
            a = t[slots[i]]
            b = t[slots[n - 1 - i]]
            if "Bye" in (a, b):
                real = b if a == "Bye" else a
                pairings.append(("Bye", real))
            else:
                # Assign home to the team with fewer home games; tie goes to a
                if home_count[a] <= home_count[b]:
                    pairings.append((b, a))   # away=b, home=a
                    home_count[a] += 1
                else:
                    pairings.append((a, b))   # away=a, home=b
                    home_count[b] += 1
        rounds.append(pairings)

        # Clockwise rotation: pin slots[0], move last slot to position 1
        slots = [slots[0]] + [slots[-1]] + slots[1:-1]

    return rounds


# ── Date helpers ──────────────────────────────────────────────────────────────

def _parse_time(s: str) -> time:
    h, m = map(int, s.strip().split(":"))
    return time(h, m)


def _add_hours(t: time, hours: float) -> time:
    total = t.hour * 60 + t.minute + int(hours * 60)
    return time((total // 60) % 24, total % 60)


def _next_valid_date(d: date, weekdays: set) -> date:
    while d.weekday() not in weekdays:
        d += timedelta(days=1)
    return d


# ── Generate view ─────────────────────────────────────────────────────────────

class ScheduleGenerateView(APIView):
    """
    POST /api/schedules/generate/

    Body:
      teams            [str]   – team names
      event_type       str     – "GAME" | "PRACTICE" | "OTHER"  (default "GAME")
      rounds           int     – total rounds (GAME only; default = n_teams - 1)
      start_date       str     – "YYYY-MM-DD" (optional)
      game_days        [int]   – weekday numbers 0=Mon…6=Sun
      time_slots       [str]   – HH:MM start times per day (one per group)
      duration_hours   float   – event length in hours (default 2)
      location         str
      field            str
      teams_per_day    int     – PRACTICE only: how many teams share each day,
                                 stacked consecutively (default 1).
                                 E.g. 4 = all 4 teams back-to-back on one day.
    """

    authentication_classes = []
    permission_classes = []

    def post(self, request):
        teams = request.data.get("teams", [])
        if not teams or len(teams) < 1:
            return Response({"error": "Provide at least 1 team."}, status=400)

        event_type     = (request.data.get("event_type") or "GAME").upper()
        start_date_str = request.data.get("start_date", "")
        game_days_raw  = request.data.get("game_days", [])
        time_slots_raw = request.data.get("time_slots", ["18:00"])
        duration_hours = float(request.data.get("duration_hours", 2))
        location       = request.data.get("location", "")
        field          = request.data.get("field", "")

        use_dates = bool(start_date_str and game_days_raw and time_slots_raw)
        game_days = set(int(d) for d in game_days_raw) if use_dates else set()
        time_slots = [s.strip() for s in time_slots_raw if s.strip()] if use_dates else ["18:00"]

        games = []

        if event_type == "GAME":
            if len(teams) < 2:
                return Response({"error": "Games require at least 2 teams."}, status=400)

            t_padded = list(teams) + (["Bye"] if len(teams) % 2 == 1 else [])
            n_base_rounds = len(t_padded) - 1
            num_rounds = int(request.data.get("rounds", n_base_rounds))
            base_rounds = _generate_base_rounds(teams)
            all_rounds = [base_rounds[i % len(base_rounds)] for i in range(num_rounds)]

            order = 1
            if use_dates:
                current_date = _next_valid_date(
                    datetime.strptime(start_date_str, "%Y-%m-%d").date(), game_days
                )
                slot_idx = 0

                def consume_slot():
                    nonlocal current_date, slot_idx
                    t_str = time_slots[slot_idx]
                    used_date = current_date
                    slot_idx += 1
                    if slot_idx >= len(time_slots):
                        slot_idx = 0
                        current_date = _next_valid_date(current_date + timedelta(days=1), game_days)
                    return used_date, t_str

                for r_idx, pairings in enumerate(all_rounds):
                    round_num = r_idx + 1
                    real = [(a, h) for a, h in pairings if a != "Bye"]
                    byes = [(a, h) for a, h in pairings if a == "Bye"]
                    round_date = None

                    for away, home in real:
                        g_date, t_str = consume_slot()
                        if round_date is None:
                            round_date = g_date
                        start_t = _parse_time(t_str)
                        end_t = _add_hours(start_t, duration_hours)
                        games.append({
                            "order": order, "round": round_num,
                            "away_team": away, "home_team": home,
                            "date": g_date.strftime("%m/%d/%Y"),
                            "start_time": start_t.strftime("%H:%M"),
                            "end_time": end_t.strftime("%H:%M"),
                            "location": location, "field": field,
                        })
                        order += 1

                    for away, home in byes:
                        games.append({
                            "order": order, "round": round_num,
                            "away_team": away, "home_team": home,
                            "date": round_date.strftime("%m/%d/%Y") if round_date else "",
                            "start_time": None, "end_time": None,
                            "location": None, "field": None,
                        })
                        order += 1
            else:
                for r_idx, pairings in enumerate(all_rounds):
                    round_num = r_idx + 1
                    for away, home in pairings:
                        is_bye = away == "Bye"
                        games.append({
                            "order": order, "round": round_num,
                            "away_team": away, "home_team": home,
                            "date": "", "start_time": None, "end_time": None,
                            "location": None if is_bye else location,
                            "field": None if is_bye else field,
                        })
                        order += 1

        else:
            # ── PRACTICE / OTHER ──────────────────────────────────────────────
            #
            # teams_per_day: how many teams share one calendar day.
            #   = 1  (default): each team gets its own day at time_slots[0]
            #   > 1: teams are stacked consecutively — team N+1 starts when
            #        team N ends, all on the same date.
            #
            # Example (Majors, teams_per_day=4, start=10:00, dur=1.5h):
            #   Tigers   10:00–11:30
            #   Guardians 11:30–13:00
            #   Royals   13:00–14:30
            #   Twins    14:30–16:00   ← all on the same Saturday

            label = "Practice" if event_type == "PRACTICE" else "Other"
            num_rounds = int(request.data.get("rounds", 1))
            teams_per_day = max(1, int(request.data.get("teams_per_day", 1)))
            order = 1

            if use_dates:
                current_date = _next_valid_date(
                    datetime.strptime(start_date_str, "%Y-%m-%d").date(), game_days
                )
                slot_idx = 0   # indexes into time_slots for the group's base start time

                for r_idx in range(num_rounds):
                    round_num = r_idx + 1

                    # Divide teams into groups; each group shares one day
                    for g_start in range(0, len(teams), teams_per_day):
                        group = teams[g_start : g_start + teams_per_day]
                        day = current_date

                        # Base start time for this group
                        base_str = time_slots[slot_idx % len(time_slots)]
                        slot_time = _parse_time(base_str)

                        for team in group:
                            end_t = _add_hours(slot_time, duration_hours)
                            games.append({
                                "order": order, "round": round_num,
                                "away_team": label, "home_team": team,
                                "date": day.strftime("%m/%d/%Y"),
                                "start_time": slot_time.strftime("%H:%M"),
                                "end_time": end_t.strftime("%H:%M"),
                                "location": location, "field": field,
                            })
                            order += 1
                            # Next team starts when this one ends (consecutive stack)
                            slot_time = end_t

                        # Advance to next valid day and next time slot for next group
                        slot_idx += 1
                        current_date = _next_valid_date(current_date + timedelta(days=1), game_days)
            else:
                # No dates — generate order only, no stacking logic needed
                for r_idx in range(num_rounds):
                    round_num = r_idx + 1
                    for g_start in range(0, len(teams), teams_per_day):
                        group = teams[g_start : g_start + teams_per_day]
                        for team in group:
                            games.append({
                                "order": order, "round": round_num,
                                "away_team": label, "home_team": team,
                                "date": "", "start_time": None, "end_time": None,
                                "location": location, "field": field,
                            })
                            order += 1

        return Response({"games": games, "total_rounds": max((g["round"] for g in games), default=0), "teams": teams})


# ── Export view ───────────────────────────────────────────────────────────────

class ScheduleExportView(APIView):
    """
    POST /api/schedules/export/

    Body:
      games  [{order, round, away_team, home_team, date, start_time,
               end_time, location, field}]
      title  str  – used as filename and sheet note
    """

    authentication_classes = []
    permission_classes = []

    def post(self, request):
        games = request.data.get("games", [])
        title = request.data.get("title", "Game Schedule")

        # Detect practice/other from the games list
        practice_labels = {"Practice", "Other"}
        is_practice = (
            bool(games) and
            all(g.get("away_team") in practice_labels for g in games if g.get("away_team") != "Bye")
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # ── Header — SportsConnect format differs for practice vs game ────────
        if is_practice:
            # SportsConnect practice format: SortOrder RoundNo Team MatchDate StartTime EndTime Location Field
            headers = ["SortOrder", "RoundNo", "Team",
                       "MatchDate", "StartTime", "EndTime", "Location", "Field"]
        else:
            headers = ["Order", "Round", "Away Team", "Home Team",
                       "Date", "Start Time", "End Time", "Location", "Field"]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1F497D")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 18

        # ── Data rows ────────────────────────────────────────────────────────
        center = Alignment(horizontal="center", vertical="center")

        for game in games:
            if is_practice:
                row = [
                    game.get("order"),
                    game.get("round"),
                    game.get("home_team"),       # Team (home_team = the actual team)
                    game.get("date") or None,
                    game.get("start_time") or None,
                    game.get("end_time") or None,
                    game.get("location") or None,
                    game.get("field") or None,
                ]
                ws.append(row)
                r = ws.max_row
                for col in range(1, 9):
                    cell = ws.cell(r, col)
                    cell.alignment = center if col in (1, 2, 4, 5, 6) else Alignment(vertical="center")
                    cell.font = Font(name="Calibri", size=10)
            else:
                is_bye = game.get("away_team") == "Bye"
                bye_fill = PatternFill("solid", fgColor="F2F2F2")
                row = [
                    game.get("order"),
                    game.get("round"),
                    game.get("away_team"),
                    game.get("home_team"),
                    game.get("date") or None,
                    game.get("start_time") or None,
                    game.get("end_time") or None,
                    game.get("location") or None,
                    game.get("field") or None,
                ]
                ws.append(row)
                r = ws.max_row
                for col in range(1, 10):
                    cell = ws.cell(r, col)
                    cell.alignment = center if col in (1, 2, 5, 6, 7) else Alignment(vertical="center")
                    if is_bye:
                        cell.fill = bye_fill
                        cell.font = Font(color="888888", name="Calibri", size=10, italic=True)
                    else:
                        cell.font = Font(name="Calibri", size=10)

        # ── Column widths ────────────────────────────────────────────────────
        if is_practice:
            widths = [10, 8, 28, 12, 11, 11, 16, 14]
        else:
            widths = [7, 7, 26, 26, 12, 11, 11, 16, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Freeze header ────────────────────────────────────────────────────
        ws.freeze_panes = "A2"

        # ── Stream response ──────────────────────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe = title.replace(" ", "_").replace("/", "-")
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{safe}.xlsx"'
        return resp
